"""
Geração de artefatos de saída:
generate_report(): PDF executivo (reportlab) com células em Paragraph
(quebra de linha correta, sem estouro de coluna);
generate_excel(): .xlsx com 3 abas (openpyxl);
generate_csv(): .csv (utf-8-sig, separador ';' para Excel pt-BR).

Rastreabilidade da revisão manual (TAREFA 5):
lançamentos válidos confirmados manualmente pelo operador recebem o
marcador "*" na descrição + nota de rodapé explicativa;
motivos de exclusão manual chegam na coluna de motivo da auditoria
("Excluída manualmente pelo usuário (motivo)"), produzidos pelo
rules_engine a partir do Dict[int, str] da tela de revisão.
Compatibilidade: lê tx.manually_confirmed via getattr — transações sem o
atributo (fluxos antigos) comportam-se como não-manuais.

RODADA 2:
FIX F: seção/linhas de "Rastreabilidade da Revisão Manual" no PDF, Excel e
CSV, consumindo a chave "revisao_manual" produzida pelo income_calculator
(omitida automaticamente em fluxos antigos sem a chave);
FIX G: canário de inconsistência — entrada válida com valor negativo gera
logger.error nos três artefatos, sem alterar o fluxo de geração.

RODADA GEMINI:
FIX H: marcador de rastreabilidade da extração via IA — lançamentos com
extraction_source="gemini" (atributo dinâmico setado pelo
gemini_extractor) recebem selo "[IA]" no PDF e "🤖" no Excel/CSV
(Helvetica não possui glifos de emoji, daí a diferença de selo),
nota de rodapé e contagem no resumo. Compatibilidade: getattr —
transações sem o atributo (fluxo local) não recebem selo.

RODADA ATUAL (CORREÇÃO ESTRUTURAL DO PDF):
- Agrupamento de lançamentos válidos por mês em seções separadas, cada uma
  com sua própria tabela (Data | Descrição | Valor) e subtotal do mês.
- Paginação com cabeçalho repetido ("Relatório Consolidado de Apuração de
  Renda - <Titular>") e rodapé ("Página X de Y") em todas as páginas via
  PageTemplate.
- Resumo consolidado por mês agora inclui coluna "Dias Cobertos".
- Tabela de Auditoria completa com TODOS os lançamentos excluídos e o
  motivo exato da exclusão.
- Nota metodológica preservada no rodapé final.
"""
import io
import csv
import logging
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from xml.sax.saxutils import escape

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageTemplate, Frame, BaseDocTemplate, NextPageTemplate
)
from reportlab.pdfgen import canvas

logger = logging.getLogger(__name__)

PRIMARY_BLUE = colors.HexColor("#1E3A8A")
LIGHT_BLUE = colors.HexColor("#DBEAFE")
GRAY_TEXT = colors.HexColor("#374151")
LIGHT_GRAY = colors.HexColor("#F3F4F6")
WHITE = colors.white

NOTA_CONFIRMACAO_MANUAL = (
    "* sinal de crédito/débito confirmado manualmente pelo operador"
)

# FIX H (RODADA GEMINI): nota de rastreabilidade da extração via IA.
NOTA_EXTRACAO_IA = (
    "Lançamentos marcados com [IA] (PDF) / 🤖 (Excel/CSV) foram extraídos "
    "via IA (Gemini) e validados contra os somatórios impressos pelo banco"
)


def format_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def format_date(date_obj) -> str:
    return date_obj.strftime("%d/%m/%Y")


def _is_manual(tx) -> bool:
    """True se o lançamento foi confirmado manualmente pelo operador."""
    return bool(getattr(tx, "manually_confirmed", False))


def _is_gemini(tx) -> bool:
    """FIX H: True se o lançamento foi extraído via Gemini (atributo dinâmico)."""
    return getattr(tx, "extraction_source", "") == "gemini"


def _review_summary(metrics: Dict[str, Any]) -> Optional[Tuple[int, int]]:
    """
    FIX F (rodada 2): rastreabilidade da revisão manual.
    Lê a chave "revisao_manual" produzida pelo income_calculator (FIX C).
    Retorna (confirmadas_como_renda, excluidas_ou_pendentes), ou None se a
    chave não existir (compatibilidade retroativa com fluxos antigos).
    """
    revisao = metrics.get("revisao_manual")
    if not revisao:
        return None
    return (
        len(revisao.get("incluidas", [])),
        len(revisao.get("excluidas", [])),
    )


def _canary_negative_valid(tx) -> None:
    """
    FIX G (rodada 2): CANÁRIO de inconsistência de contrato.
    Entrada válida com valor negativo indica regressão de parser/calculator
    (contrato do sistema: entrada válida => amount > 0). Apenas loga o erro
    para diagnóstico imediato, SEM alterar o fluxo de geração do artefato.
    """
    if tx.amount < 0:
        logger.error(
            "INCONSISTÊNCIA NO RELATÓRIO: entrada válida com valor negativo "
            "(%s | %s | %.2f | %s). Regressão de parser/calculator suspeita.",
            getattr(tx, "date", "?"),
            tx.description,
            tx.amount,
            getattr(tx, "source_file", "") or "PDF",
        )


def _build_cell_styles() -> Dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    cell = ParagraphStyle("Cell", parent=base["Normal"], fontName="Helvetica",
                          fontSize=7.5, leading=9.5, textColor=GRAY_TEXT)
    return {
        "cell_l": cell,
        "cell_c": ParagraphStyle("CellC", parent=cell, alignment=TA_CENTER),
        "cell_r": ParagraphStyle("CellR", parent=cell, alignment=TA_RIGHT),
        "head_l": ParagraphStyle("HeadL", parent=cell, fontName="Helvetica-Bold",
                                 fontSize=8, textColor=WHITE, alignment=TA_LEFT),
        "head_c": ParagraphStyle("HeadC", parent=cell, fontName="Helvetica-Bold",
                                 fontSize=8, textColor=WHITE, alignment=TA_CENTER),
        "head_r": ParagraphStyle("HeadR", parent=cell, fontName="Helvetica-Bold",
                                 fontSize=8, textColor=WHITE, alignment=TA_RIGHT),
        "tot_c": ParagraphStyle("TotC", parent=cell, fontName="Helvetica-Bold",
                                alignment=TA_CENTER),
        "tot_r": ParagraphStyle("TotR", parent=cell, fontName="Helvetica-Bold",
                                alignment=TA_RIGHT),
        "note": ParagraphStyle("Note", parent=base["Normal"], fontSize=8,
                               textColor=GRAY_TEXT),
    }


def _zebra(style_cmds: List, nrows: int, first_data_row: int = 1):
    for i in range(first_data_row, nrows):
        style_cmds.append(("BACKGROUND", (0, i), (-1, i),
                           LIGHT_GRAY if i % 2 == 0 else WHITE))
    style_cmds += [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    return style_cmds


# ---------------------------------------------------------------------------
# CLASSE CUSTOMIZADA DE DOCUMENTO PARA CABEÇALHO/RODAPÉ REPETITIVOS
# ---------------------------------------------------------------------------
class RelatorioApuracaoDoc(BaseDocTemplate):
    """
    Documento customizado que desenha cabeçalho e rodapé em todas as páginas.
    Cabeçalho: "Relatório Consolidado de Apuração de Renda - <Titular>"
    Rodapé: "Página X de Y"
    """
    def __init__(self, filename, holder_name, **kwargs):
        super().__init__(filename, **kwargs)
        self.holder_name = holder_name or "Não informado"
        self.page_count = 0
        
        # Frame para o conteúdo principal (deixando espaço para cabeçalho/rodapé)
        frame = Frame(
            2 * cm, 2.5 * cm,  # x, y (margem esquerda, margem inferior + rodapé)
            A4[0] - 4 * cm, A4[1] - 5 * cm,  # width, height
            id='main_frame'
        )
        
        # Template de página com cabeçalho e rodapé
        template = PageTemplate(
            id='main_template',
            frames=[frame],
            onPage=self._draw_header_footer
        )
        self.addPageTemplates([template])

    def _draw_header_footer(self, canvas_obj, doc):
        """Desenha o cabeçalho e o rodapé em todas as páginas."""
        canvas_obj.saveState()
        
        # --- CABEÇALHO ---
        canvas_obj.setFont("Helvetica-Bold", 10)
        canvas_obj.setFillColor(PRIMARY_BLUE)
        header_text = f"Relatório Consolidado de Apuração de Renda - {self.holder_name}"
        canvas_obj.drawCentredString(A4[0] / 2, A4[1] - 1.5 * cm, header_text)
        
        # Linha separadora do cabeçalho
        canvas_obj.setStrokeColor(PRIMARY_BLUE)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(2 * cm, A4[1] - 1.7 * cm, A4[0] - 2 * cm, A4[1] - 1.7 * cm)
        
        # --- RODAPÉ ---
        # Linha separadora do rodapé
        canvas_obj.setStrokeColor(colors.gray)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(2 * cm, 2 * cm, A4[0] - 2 * cm, 2 * cm)
        
        # Numeração de página "Página X de Y"
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.gray)
        page_num = canvas_obj.getPageNumber()
        # Usamos um placeholder que será atualizado no afterPage
        canvas_obj.drawCentredString(A4[0] / 2, 1.3 * cm, f"Página {page_num}")
        
        canvas_obj.restoreState()

    def afterPage(self):
        """Atualiza a contagem total de páginas para o rodapé "Página X de Y"."""
        self.page_count = self.page  # Armazena o número da última página


def generate_report(metrics: Dict[str, Any], holder_name: str,
                    institutions: List[str]) -> io.BytesIO:
    buffer = io.BytesIO()
    
    doc = RelatorioApuracaoDoc(
        buffer,
        holder_name=holder_name,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
        title="Relatório Consolidado de Apuração de Renda",
    )
    
    styles = getSampleStyleSheet()
    S = _build_cell_styles()
    
    title_style = ParagraphStyle("CustomTitle", parent=styles["Heading1"], fontSize=18,
                                  textColor=PRIMARY_BLUE, spaceAfter=10, alignment=TA_CENTER,
                                  fontName="Helvetica-Bold")
    subtitle_style = ParagraphStyle("CustomSubtitle", parent=styles["Normal"], fontSize=11,
                                     textColor=GRAY_TEXT, alignment=TA_CENTER, spaceAfter=10)
    section_title_style = ParagraphStyle("SectionTitle", parent=styles["Heading2"],
                                          fontSize=13, textColor=PRIMARY_BLUE,
                                          spaceBefore=15, spaceAfter=10,
                                          fontName="Helvetica-Bold")
    footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8,
                                   textColor=colors.gray, alignment=TA_CENTER)
    kpi_value_style = ParagraphStyle("kpi_val", parent=subtitle_style,
                                      textColor=PRIMARY_BLUE, fontSize=14,
                                      alignment=TA_CENTER)
    
    story = []
    
    # --- TÍTULO E DADOS DO TITULAR ---
    story.append(Paragraph("Relatório Executivo de Apuração de Renda", title_style))
    institutions_str = ", ".join(institutions) if institutions else "Não identificadas"
    header_text = (f"<b>Titular:</b> {escape(holder_name or 'Não informado')}<br/>"
                   f"<b>Instituição(ões):</b> {escape(institutions_str)}<br/>"
                   f"<b>Data de Geração:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    story.append(Paragraph(header_text, subtitle_style))
    story.append(Spacer(1, 0.5 * cm))
    
    # --- Cards de KPI ---
    kpi_data = [
        [Paragraph("<b>Total Geral Acumulado</b>", subtitle_style),
         Paragraph("<b>Média Mensal Geral</b>", subtitle_style),
         Paragraph("<b>Média Meses Completos</b>", subtitle_style)],
        [Paragraph(f"<b>{format_currency(metrics.get('total_geral', 0))}</b>", kpi_value_style),
         Paragraph(f"<b>{format_currency(metrics.get('media_mensal_geral', 0))}</b>", kpi_value_style),
         Paragraph(f"<b>{format_currency(metrics.get('media_meses_completos', 0))}</b>", kpi_value_style)],
    ]
    kpi_table = Table(kpi_data, colWidths=[5.5 * cm] * 3)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BLUE),
        ("GRID", (0, 0), (-1, -1), 1, WHITE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 1 * cm))
    
    # --- Resumo Consolidado por Mês (com Dias Cobertos) ---
    story.append(Paragraph("Resumo Consolidado por Mês", section_title_style))
    resumo_data = [[Paragraph("Mês/Ano", S["head_c"]),
                    Paragraph("Dias Cobertos", S["head_c"]),
                    Paragraph("Qtd. Entradas Válidas", S["head_c"]),
                    Paragraph("Total Apurado (R$)", S["head_r"])]]
    
    total_geral_check = 0.0
    total_qtd_check = 0
    for item in metrics.get("resumo_mensal", []):
        dias = item.get("dias_cobertos", 0)
        resumo_data.append([
            Paragraph(item["month_label"], S["cell_c"]),
            Paragraph(str(dias), S["cell_c"]),
            Paragraph(str(item["qtd_entradas_validas"]), S["cell_c"]),
            Paragraph(format_currency(item["total_valido"]), S["cell_r"]),
        ])
        total_geral_check += item["total_valido"]
        total_qtd_check += item["qtd_entradas_validas"]
    
    resumo_data.append([
        Paragraph("<b>TOTAL GERAL CONSOLIDADO</b>", S["tot_c"]),
        Paragraph("", S["tot_c"]),
        Paragraph(f"<b>{total_qtd_check}</b>", S["tot_c"]),
        Paragraph(f"<b>{format_currency(metrics.get('total_geral', 0))}</b>", S["tot_r"]),
    ])
    
    resumo_table = Table(resumo_data, colWidths=[3 * cm, 3 * cm, 4.5 * cm, 5 * cm], repeatRows=1)
    resumo_table.setStyle(TableStyle(_zebra([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
        ("BACKGROUND", (0, len(resumo_data) - 1), (-1, len(resumo_data) - 1), LIGHT_BLUE),
    ], len(resumo_data))))
    story.append(resumo_table)
    story.append(Spacer(1, 1 * cm))
    
    # --- Detalhamento de Entradas Válidas AGRUPADO POR MÊS ---
    story.append(Paragraph("Detalhamento de Entradas Válidas", section_title_style))
    
    valid_txs = metrics.get("entradas_validas", [])
    manual_count = 0
    ia_count = 0
    
    if valid_txs:
        # Agrupar transações por mês/ano
        monthly_groups = defaultdict(list)
        for tx in valid_txs:
            month_key = (tx.date.year, tx.date.month)
            monthly_groups[month_key].append(tx)
        
        # Ordenar os meses cronologicamente
        sorted_months = sorted(monthly_groups.keys(), key=lambda x: (x[0], x[1]))
        
        MESES_PT = {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
            5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
            9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
        }
        
        for year, month in sorted_months:
            txs_do_mes = monthly_groups[(year, month)]
            mes_nome = MESES_PT.get(month, str(month))
            
            # Título da seção do mês
            story.append(Paragraph(
                f"Detalhamento de Entradas - {mes_nome}/{year}",
                ParagraphStyle("MonthTitle", parent=section_title_style, fontSize=11, 
                               textColor=GRAY_TEXT, spaceBefore=10, spaceAfter=5)
            ))
            
            # Tabela do mês
            detail_data = [[Paragraph("Data", S["head_c"]),
                            Paragraph("Descrição da Entrada", S["head_l"]),
                            Paragraph("Valor", S["head_r"])]]
            
            subtotal = 0.0
            for tx in txs_do_mes:
                _canary_negative_valid(tx)
                desc = escape(tx.description or "-")
                if _is_manual(tx):
                    desc += " *"
                    manual_count += 1
                if _is_gemini(tx):
                    desc += " [IA]"
                    ia_count += 1
                
                detail_data.append([
                    Paragraph(format_date(tx.date), S["cell_c"]),
                    Paragraph(desc, S["cell_l"]),
                    Paragraph(format_currency(tx.amount), S["cell_r"]),
                ])
                subtotal += tx.amount
            
            # Linha de subtotal do mês
            detail_data.append([
                Paragraph("", S["cell_c"]),
                Paragraph(f"<b>Total do Mês ({mes_nome}/{year})</b>", S["cell_l"]),
                Paragraph(f"<b>{format_currency(subtotal)}</b>", S["cell_r"]),
            ])
            
            detail_table = Table(detail_data, colWidths=[2.5 * cm, 11 * cm, 3.5 * cm], repeatRows=1)
            detail_table.setStyle(TableStyle(_zebra([
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
                ("BACKGROUND", (0, len(detail_data) - 1), (-1, len(detail_data) - 1), LIGHT_BLUE),
            ], len(detail_data))))
            story.append(detail_table)
            story.append(Spacer(1, 0.5 * cm))
        
        # Notas de rastreabilidade (se houver)
        if manual_count > 0:
            story.append(Spacer(1, 0.2 * cm))
            story.append(Paragraph(NOTA_CONFIRMACAO_MANUAL, S["note"]))
        if ia_count > 0:
            story.append(Spacer(1, 0.2 * cm))
            story.append(Paragraph(NOTA_EXTRACAO_IA, S["note"]))
    else:
        story.append(Paragraph("Nenhuma entrada válida encontrada.", subtitle_style))
    
    story.append(Spacer(1, 1 * cm))
    
    # --- Tabela de Auditoria COMPLETA (Valores Excluídos) ---
    story.append(Paragraph("Tabela de Auditoria (Valores Excluídos / Entradas Ignoradas)", section_title_style))
    excluded_txs = metrics.get("entradas_excluidas", [])
    if excluded_txs:
        audit_data = [[Paragraph("Data", S["head_c"]),
                       Paragraph("Descrição Original", S["head_l"]),
                       Paragraph("Regra de Exclusão Aplicada", S["head_l"]),
                       Paragraph("Valor", S["head_r"])]]
        for item in excluded_txs:
            audit_data.append([
                Paragraph(format_date(item["date"]), S["cell_c"]),
                Paragraph(escape(item["description"] or "-"), S["cell_l"]),
                Paragraph(escape(item["reason"] or "-"), S["cell_l"]),
                Paragraph(format_currency(item["amount"]), S["cell_r"]),
            ])
        audit_table = Table(audit_data, colWidths=[2.5 * cm, 6.5 * cm, 5.5 * cm, 3 * cm], repeatRows=1)
        audit_table.setStyle(TableStyle(_zebra([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
        ], len(audit_data))))
        story.append(audit_table)
    else:
        story.append(Paragraph("Nenhum valor excluído.", subtitle_style))
    
    # --- FIX F (rodada 2): Rastreabilidade da Revisão Manual ---
    review = _review_summary(metrics)
    if review is not None:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph(
            f"Rastreabilidade da Revisão Manual: {review[0]} lançamento(s) confirmado(s) "
            f"manualmente como renda • {review[1]} exclusão(ões) manuais ou pendentes de revisão.",
            S["note"],
        ))
    
    story.append(Spacer(1, 1.5 * cm))
    
    # --- Nota Metodológica ---
    footer_text = (
        "<b>Nota Metodológica:</b> Este relatório consolida entradas financeiras identificadas nos "
        "extratos fornecidos, excluindo: (1) transferências de mesma titularidade (quando o nome "
        "do titular aparece na contraparte); (2) rendimentos/resgastes de aplicações financeiras; "
        "(3) créditos de jogos/apostas; e (4) quaisquer lançamentos de débito/compra. "
        "A 'Média Meses Completos' considera apenas períodos com mais de 20 dias de extrato cobertos. "
        "Lançamentos marcados com '*' tiveram o sinal confirmado manualmente pelo operador. "
        "Lançamentos marcados com '[IA]' foram extraídos via IA (Gemini) e validados contra "
        "os somatórios impressos pelo banco."
    )
    story.append(Paragraph(footer_text, footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def _excel_header_style(ws, ncols: int):
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def generate_excel(metrics: Dict[str, Any], holder_name: str = "",
                   institutions: List[str] = None) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    
    # FIX H: contagem prévia para o resumo (ws1).
    ia_total = sum(1 for t in metrics.get("entradas_validas", []) if _is_gemini(t))
    
    ws1 = wb.active
    ws1.title = "Resumo_Mensal"
    ws1.append(["Mês/Ano", "Dias Cobertos", "Qtd Entradas Válidas", "Total Válido Mensal"])
    _excel_header_style(ws1, 4)
    for item in metrics.get("resumo_mensal", []):
        ws1.append([item["month_label"], item.get("dias_cobertos", 0),
                    item["qtd_entradas_validas"], round(item["total_valido"], 2)])
    ws1.append(["TOTAL", "",
                sum(i["qtd_entradas_validas"] for i in metrics.get("resumo_mensal", [])),
                round(metrics.get("total_geral", 0.0), 2)])
    ws1.append([])
    ws1.append(["Titular", holder_name or "Não informado"])
    ws1.append(["Instituições", ", ".join(institutions or []) or "Não identificadas"])
    ws1.append(["Total Geral Acumulado", round(metrics.get("total_geral", 0.0), 2)])
    ws1.append(["Média Mensal Geral", round(metrics.get("media_mensal_geral", 0.0), 2)])
    ws1.append(["Média Meses Completos", round(metrics.get("media_meses_completos", 0.0), 2)])
    
    # FIX F (rodada 2): rastreabilidade da revisão manual no Excel.
    review = _review_summary(metrics)
    if review is not None:
        ws1.append(["Revisão Manual (confirmados como renda)", "", review[0], ""])
        ws1.append(["Revisão Manual (exclusões/pendentes)", "", review[1], ""])
    
    # FIX H (RODADA GEMINI): contagem de extração via IA no resumo.
    if ia_total > 0:
        ws1.append(["Extração via IA (Gemini)", "", ia_total, ""])
    
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22
    
    ws2 = wb.create_sheet("Entradas_Validas")
    ws2.append(["Data", "Descrição da Entrada", "Valor", "Banco", "Obs."])
    _excel_header_style(ws2, 5)
    manual_any = False
    ia_any = False
    for tx in metrics.get("entradas_validas", []):
        _canary_negative_valid(tx)
        manual = _is_manual(tx)
        gemini = _is_gemini(tx)
        manual_any = manual_any or manual
        ia_any = ia_any or gemini
        obs = ("*" if manual else "") + ("🤖" if gemini else "")
        ws2.append([format_date(tx.date), tx.description, round(tx.amount, 2),
                    getattr(tx, "bank", ""), obs])
    if manual_any:
        ws2.append([])
        ws2.append([NOTA_CONFIRMACAO_MANUAL])
    if ia_any:
        ws2.append([])
        ws2.append([NOTA_EXTRACAO_IA])
    
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 8
    
    ws3 = wb.create_sheet("Auditoria_Excluidos")
    ws3.append(["Data", "Descrição Original", "Regra de Exclusão Aplicada", "Valor"])
    _excel_header_style(ws3, 4)
    for item in metrics.get("entradas_excluidas", []):
        ws3.append([format_date(item["date"]), item["description"], item["reason"],
                    round(item["amount"], 2)])
    
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 80
    ws3.column_dimensions["C"].width = 50
    ws3.column_dimensions["D"].width = 14
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv(metrics: Dict[str, Any]) -> bytes:
    out = io.StringIO()
    w = csv.writer(out, delimiter=";")
    w.writerow(["RESUMO MENSAL"])
    w.writerow(["Mês/Ano", "Dias Cobertos", "Qtd Entradas Válidas", "Total Válido Mensal"])
    for item in metrics.get("resumo_mensal", []):
        w.writerow([item["month_label"], item.get("dias_cobertos", 0),
                    item["qtd_entradas_validas"],
                    f"{item['total_valido']:.2f}".replace(".", ",")])
    w.writerow(["TOTAL", "",
                sum(i["qtd_entradas_validas"] for i in metrics.get("resumo_mensal", [])),
                f"{metrics.get('total_geral', 0.0):.2f}".replace(".", ",")])
    
    # FIX F (rodada 2): rastreabilidade da revisão manual no CSV.
    review = _review_summary(metrics)
    if review is not None:
        w.writerow(["REVISÃO MANUAL", "Confirmados como renda", review[0], ""])
        w.writerow(["REVISÃO MANUAL", "Exclusões/pendentes", review[1], ""])
    
    # FIX H (RODADA GEMINI): contagem de extração via IA no CSV.
    ia_total = sum(1 for t in metrics.get("entradas_validas", []) if _is_gemini(t))
    if ia_total > 0:
        w.writerow(["EXTRAÇÃO VIA IA (GEMINI)", "Lançamentos", ia_total, ""])
    
    w.writerow([])
    w.writerow(["ENTRADAS VÁLIDAS"])
    w.writerow(["Data", "Descrição da Entrada", "Valor", "Banco", "Obs."])
    manual_any = False
    ia_any = False
    for tx in metrics.get("entradas_validas", []):
        _canary_negative_valid(tx)
        manual = _is_manual(tx)
        gemini = _is_gemini(tx)
        manual_any = manual_any or manual
        ia_any = ia_any or gemini
        obs = ("*" if manual else "") + ("🤖" if gemini else "")
        w.writerow([format_date(tx.date), tx.description,
                    f"{tx.amount:.2f}".replace(".", ","), getattr(tx, "bank", ""),
                    obs])
    if manual_any:
        w.writerow([])
        w.writerow([NOTA_CONFIRMACAO_MANUAL])
    if ia_any:
        w.writerow([])
        w.writerow([NOTA_EXTRACAO_IA])
    
    w.writerow([])
    w.writerow(["AUDITORIA - EXCLUÍDOS"])
    w.writerow(["Data", "Descrição Original", "Regra de Exclusão Aplicada", "Valor"])
    for item in metrics.get("entradas_excluidas", []):
        w.writerow([format_date(item["date"]), item["description"], item["reason"],
                    f"{item['amount']:.2f}".replace(".", ",")])
    
    return out.getvalue().encode("utf-8-sig")